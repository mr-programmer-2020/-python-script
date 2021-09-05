from kivy.app import App   
from kivy.lang import Builder
from kivy.uix.screenmanager import Screen
from kivy.uix.button import ButtonBehavior, Button
from kivy.uix.floatlayout import FloatLayout
from kivy.uix.image import Image
from kivy.animation import Animation
from kivy.uix.widget import Widget
from kivy_garden.graph import Graph, MeshLinePlot
from kivy.properties import ObjectProperty
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from kivy.uix.gridlayout import GridLayout
from kivy.app import runTouchApp
from email import encoders
from openpyxl import Workbook
from openpyxl import load_workbook
from math import sin
import xlsxwriter



class MyLayout_W(Widget):
    pass
class FirstScrren(Screen):
    
    pass
class BarGraph(Screen):
    
 
    pass
 

class Pulse(Screen):
    i = str('15')
    pass
class Breese(Screen):
    pass
class Resalt(Screen):
    pass
class SecondScreen(Screen):
	pass

class ImageButton(ButtonBehavior, Image):

	pass

class ThirdScreen(Screen):
	pass
class Information(Screen):
    pass

class Information_veiw(Screen):
    pass
class Drags(Screen):
    pass
class DragsChoise(Screen):
    pass



GUI = Builder.load_file("main.kv")


class MainApp(App):

    def build(self):
        return GUI

    def change_screen(self, screen_name):
        screen_manager = self.root.ids['screen_manager']
        screen_manager.current = screen_name

    def animation_button(self, widget, *args):
            anim = Animation(background_color=(243/255,132/255,132/255,1), duration= .3)
            anim.start(widget)
    def animation_button_back(self, widget, *args):
            anim = Animation(background_color=(243/255,52/255,52/255,1), duration= .3)
            anim.start(widget)
    def animation_button_gray(self, widget, *args):
            anim = Animation(background_color=(131/255 , 131/255, 131/255, 1), duration= .3)
            anim.start(widget)
    def animation_button_back_gray(self, widget, *args):
            anim = Animation(background_color=(225/255,200/255,200/255,1), duration= .3)
            anim.start(widget)



    def spinner_clicked(self):
        sp= self.root.ids['drags'].ids.spinner_id 
        sp2= self.root.ids['drags'].ids.spinner_id2 
        sp3= self.root.ids['drags'].ids.spinner_id3 
        pipi= ["Выбрать", ""]
        label_limit =self.root.ids.drags.ids.label_limit 
        label_limit.text= ""

        for i in range(1,176):
            pipi.append(f"{Text_info1[i]}")
        
            
        
        sp.values = map(str, pipi)
        sp2.values = map(str, Dizar)
        sp3.values = map(str, Days)





    def drags_add(self):
        global drags_list
        global drags_doz
        global drags_rate
        global counter_drags
        counter_1 = 1
        counter_2 = 1
        for pir in range (1,9):
            label = self.root.ids.dragschoise.ids[f"label_{pir}"] 
            but_r = self.root.ids.dragschoise.ids[f"btn_{pir}"]
            label.text = "" 
            but_r.pos_hint = {"top": counter_1, "center_x": 0}
            counter_1= counter_1 - 0.1
        counter_1 = 1
        for i in drags_list.values():
            label = self.root.ids.dragschoise.ids[f"label_{counter_2}"]    
            but_r = self.root.ids.dragschoise.ids[f"btn_{counter_2}"]
            but_r.pos_hint = {"top": counter_1, "center_x": .5}
            label.text = i
            counter_1 -=0.1
            counter_2 += 1
         

    def drags_remover(self, name, num):
        global drags_list
        global drags_doz
        global drags_rate
        but_r = self.root.ids.dragschoise.ids[f"label_{num}"]
        but_r = self.root.ids.dragschoise.ids[f"btn_{num}"]
        name_value = name.text
        
        desired_value = name_value
        for key, value in drags_list.items():
          if value == desired_value:
            del drags_list[key]
            del drags_doz[key]
            del drags_rate[key]
            break



        print (drags_list, drags_doz,drags_rate)
        
    def drag_save(self, name, doz, rate):
        global drags_list
        global drags_doz
        global drags_rate
        global counter_drags
        counter_drags += 1
        if counter_drags <= 8:
            
            drags_list[f"{counter_drags}"] = f"{name}"
            drags_doz[f"{counter_drags}"] = f"{doz}"
            drags_rate[f"{counter_drags}"] = f"{rate}"
        else:
            counter_drags =5
            label_limit =self.root.ids.drags.ids.label_limit 
            label_limit.text= "УДАЛИТЕ ОДНО ЛЕКАРСТВО"

    def information(self, inform):  
        Label_1 = self.root.ids.information_veiw.ids.text_info 
        Label_1.text = Text_info[int(inform)]



    def drag_save1(self):
        print (drags_list,drags_doz,drags_rate)
        
    def sever_data(self):
        print("seve_data here")



    def graph_breese(self ):
        print ("graph1")
       
        graph = ObjectProperty()

 
        plot = MeshLinePlot(color=[1, 0, 0, 1])
        #plot1 =plot.points = [(x, sin(x / 10.)) for x in range(0, 101)]
        plot.points = [(0,0), (1,1),(2,0),(3,1)]
        # Recherche des widgets
        print(plot.points)
        graph1= self.root.ids.pulse.ids.bargraph.ids.graph_1
        graph1.add_plot(plot)
    def graph_resalt(self ):
        print ("graph1")
       
        graph = ObjectProperty()

 
        plot = MeshLinePlot(color=[1, 0, 0, 1])
        #plot1 =plot.points = [(x, sin(x / 10.)) for x in range(0, 101)]
        plot.points = [(0,0), (1,1),(2,0),(3,1)]
        # Recherche des widgets
        print(plot.points)
        graph1= self.root.ids.pulse.ids.bargraph.ids.graph_1
        graph1.add_plot(plot)



#----------------------------------------------------------------------------------------------------



    def mail_senter(self):
        try:
            my_file = open("RNPC.xlsx", 'w',encoding = 'utf-8')
            my_file.write(f'data_mail\n')
            my_file.write(f'{data_list}\n')
            my_file.write(f'{block_list}\n')
            my_file.close()

            label_mail = self.root.ids.screen3.ids.label_mail
            label_mail.text = "ЧТО-ТО НЕ ТАК"

            mail_content = f'''Здравствуйте, это сообшение присланное пациентом {data_list[1]}
            
            номер пациента: {data_list[5]} 
            день рождения {data_list[2]}

            '''
            sender_address = 'ascemme@mail.ru'
            sender = 'dpd.pdg123'
            receiver_address = f'{data_list[8]}'
            message = MIMEMultipart()
            message['From'] = sender_address
            message['To'] = receiver_address
            message['Subject'] = f'прислонно пациентом {data_list[1]}'
            message.attach(MIMEText(mail_content, 'plain'))
            attach_file_name = "RNPC.xlsx"
            attach_file = open(attach_file_name, 'rb') # Open the file as binary mode
            payload = MIMEBase('application', 'octate-stream')
            payload.set_payload((attach_file).read())
            encoders.encode_base64(payload)
            payload.add_header('Content-Decomposition', 'attachment', filename=attach_file_name)
            message.attach(payload)
            session = smtplib.SMTP_SSL('smtp.mail.ru', 465) 
            session.login(sender_address, sender)
            text = message.as_string()
            session.sendmail(sender_address, receiver_address, text)
            session.quit()
            label_mail.text = "МАЙЛ ОТПРАВЛЕН"
            print('Mail Sent')
        except:
            print("mail error")

    
    def file_reader(self):
        pass

###------------------------------------------------------------------------------------------------------
    def data_save(self):
        try:
            global v_row
            pica = set(value_line)
            workbook = xlsxwriter.Workbook('RNPC.xlsx',)
            worksheet = workbook.add_worksheet()
            
            v_row = int(v_row) + 10

            for every_drags1 in Drags_line.keys():
                for every_drag in Drags_line.values():
                    if every_drag != None:
                        worksheet.write(f"{laters[2]}{every_drags1}",f"{every_drag}")
                for every_doze in Doze_line.values():
                    if every_doze != None:
                        worksheet.write(f"{laters[3]}{every_drags1}",f"{every_doze}")
                for every_rade in Rade_line.values():
                    if every_rade != None:
                        worksheet.write(f"{laters[4]}{every_drags1}",f"{every_rade}")
            for lines in range(2, int(v_row), 10):
                if lines != None:
                    print ("save_pass")
                    lines1 = lines -1
                    for i in laters:
                        if i != 14:
                            worksheet.write(f"{laters[i]}{lines1}",f"{expenses[i]}")
                            worksheet.set_column(0, i, 20)
                        else:
                            worksheet.write(f"{laters[14]}{1}",f"{v_row}")
#-------------------------------------------------------------------------------------------------------
                if Up_line == {}:
                    for time in Time_line.values():
                        if time != None:
                            for bit in Time_line.keys():
                                worksheet.write(f"{laters[1]}{bit}",f"{time}")
                    for  time in Up_line.values():
                        for  bit in Up_line.keys():
                            if time != None:
                                worksheet.write(f"{laters[5]}{bit}",f"{time}")
                    for time in Down_line.values():
                        if time != None:
                            for bit in Down_line.keys():
                                worksheet.write(f"{laters[6]}{bit}",f"{time}")
                    for time in Pulse_line.values():
                        if time != None:
                            for bit in Pulse_line.keys():
                                worksheet.write(f"{laters[7]}{bit}",f"{time}")
                    for time in Breese_line.values():
                        if time != None:
                            for bit in Breese_line.keys():
                                worksheet.write(f"{laters[8]}{bit}",f"{time}")
                    for time in WBreeze_line.values():
                        if time != None:
                            for bit in WBreeze_line.keys():
                                worksheet.write(f"{laters[9]}{bit}",f"{time}")
                    for time in Excost_line.values():
                        if time != None:
                            for bit in Excost_line.keys():
                                worksheet.write(f"{laters[10]}{bit}",f"{time}")
                    for time in Out_line.values():
                        if time != None:
                            for bit in Out_line.keys():
                                worksheet.write(f"{laters[11]}{bit}",f"{time}")
                    for time in In_lien.values():
                        if time != None:
                            for bit in In_lien.keys():
                                worksheet.write(f"{laters[12]}{bit}",f"{time}")
                    for time in M_line.values():
                        if time != None:
                            for bit in M_line.keys():
                                worksheet.write(f"{laters[13]}{bit}",f"{time}")
                    for v_lines in V_line.values():
                        drags_counter = 0
                        if v_lines != None:
                            if drags_list != {}:
                                for drags_saving in drags_list.values():
                                    drags_counter +=1
                        drags_counter2 = int(lines) + drags_counter

                        print (drags_counter2)
                        worksheet.write(f"{laters[15]}{lines}",f"{drags_counter2}")
                    lines1 = lines -1
                    print("error")
                    print (lines1,v_row )
                if int(lines1)+10== int(v_row):
                    print ("end")
                    print (block_list)
                    worksheet.write(f"{laters[5]}{lines}",f"{int(up_row)}")
                    worksheet.write(f"{laters[6]}{lines}",f"{int(douwn_row)}")
                    worksheet.write(f"{laters[7]}{lines}",f"{int(pulse_row)}")
                    worksheet.write(f"{laters[13]}{lines}",f"{int(mass_row)}")
                    worksheet.write(f"{laters[1]}{lines}",f"{int(time_row)}")
                    worksheet.write(f"{laters[8]}{lines}",f"{breeze_row}")
                    worksheet.write(f"{laters[9]}{lines}",f"{wbreeze_row}")
                    worksheet.write(f"{laters[10]}{lines}",f"{excost_row}")
                    worksheet.write(f"{laters[13]}{lines}",f"{int(mass_row)}")
                    worksheet.write(f"{laters[12]}{lines}",f"{int(in_row)}")
                    worksheet.write(f"{laters[11]}{lines}",f"{int(out_row)}")  
                    if drags_list != {}:
                        for drags_saving in drags_list.values():
                            worksheet.write(f"{laters[2]}{lines}",f"{drags_saving}")  
                        for drags_saving in drags_doz.values():
                            worksheet.write(f"{laters[3]}{lines}",f"{drags_saving}") 
                        for drags_saving in drags_rate.values():
                            worksheet.write(f"{laters[4]}{lines}",f"{drags_saving}")
                    print ("ero")


            workbook.close()
        except:
            print("sever")
            if lim == 2:
                print ("end")
                print (block_list)
                worksheet.write(f"{laters[5]}{lines}",f"{int(up_row)}")
                worksheet.write(f"{laters[6]}{lines}",f"{int(douwn_row)}")
                worksheet.write(f"{laters[7]}{lines}",f"{int(pulse_row)}")
                worksheet.write(f"{laters[13]}{lines}",f"{int(mass_row)}")
                worksheet.write(f"{laters[1]}{lines}",f"{int(time_row)}")
                worksheet.write(f"{laters[8]}{lines}",f"{breeze_row}")
                worksheet.write(f"{laters[9]}{lines}",f"{wbreeze_row}")
                worksheet.write(f"{laters[10]}{lines}",f"{excost_row}")
                worksheet.write(f"{laters[13]}{lines}",f"{int(mass_row)}")
                worksheet.write(f"{laters[12]}{lines}",f"{int(in_row)}")
                worksheet.write(f"{laters[11]}{lines}",f"{int(out_row)}")  
                if drags_list != {}:
                    for drags_saving in drags_list.values():
                        worksheet.write(f"{laters[2]}{lines}",f"{drags_saving}")  
                    for drags_saving in drags_doz.values():
                        worksheet.write(f"{laters[3]}{lines}",f"{drags_saving}") 
                    for drags_saving in drags_rate.values():
                        worksheet.write(f"{laters[4]}{lines}",f"{drags_saving}")
            workbook.close()


    def data_reader(self):
        try:
            global laters
            global Time_line
            global Drags_line
            global Doze_line
            global Rade_line
            global Up_line
            global Down_line
            global Pulse_line
            global Breese_line
            global WBreeze_line
            global Excost_line
            global Out_line
            global In_lien
            global V_line
            global M_line
            global U_line
            global value_line
            global up_row
            global douwn_row
            global pulse_row
            global mass_row
            global time_row
            global breeze_row
            global wbreeze_row
            global excost_row
            global out_row
            global in_row
            global v_row

            workbook = load_workbook(filename="RNPC.xlsx")

            sheet = workbook.active
            #sheet=sheet.title
            sheet_Lines_V= sheet["V1"].value
            v_row = sheet_Lines_V

            for t2 in range(2, int(v_row), 10):
                if t2 != None:
                    print("reading")
                    Time_line[t2]=sheet[f"{laters[1]}{t2}"].value
                    Up_line[t2]=sheet[f"{laters[5]}{t2}"].value
                    print("printing f" ,f"{laters[5]}{t2}", Up_line[t2])
                    Down_line[t2]=sheet[f"{laters[6]}{t2}"].value
                    print("printing f" ,f"{laters[6]}{t2}", Down_line[t2])
                    Pulse_line[t2]=sheet[f"{laters[7]}{t2}"].value
                    print("printing f" ,f"{laters[7]}{t2}", Pulse_line[t2])
                    Breese_line[t2]=sheet[f"{laters[8]}{t2}"].value
                    WBreeze_line[t2]=sheet[f"{laters[9]}{t2}"].value
                    Excost_line[t2]=sheet[f"{laters[10]}{t2}"].value
                    Out_line[t2]=sheet[f"{laters[11]}{t2}"].value
                    In_lien[t2]=sheet[f"{laters[12]}{t2}"].value
                    W_line[t2]=sheet[f"{laters[13]}{t2}"].value
                    U_line[t2]=sheet[f"{laters[16]}{t2}"].value
                    M_line[t2]=sheet[f"{laters[13]}{t2}"].value
                for every_w in W_line:
                    if every_w != None:
                        for t3 in range(t2, int(every_w)):
                            Drags_line[t3]=sheet[f"{laters[2]}{t2}"].value
                            Doze_line[t3]=sheet[f"{laters[3]}{t2}"].value
                            Rade_line[t3]=sheet[f"{laters[4]}{t2}"].value
            if Up_line != {}:
                for up_line in Up_line.values():
                    if up_line != None:
                        up_row = int(up_line)
                for down_line in Down_line.values():
                    if down_line != None:
                        douwn_row = int(down_line)
                for pulse_line in Pulse_line.values():
                    if pulse_line != None:
                        pulse_row = int(pulse_line)
                for time_line in Time_line.values():
                    if time_line != None:
                        time_row = int(time_line)
                for breese_line in Breese_line.values():
                    if breese_line != None:
                        print (breese_line)
                        breeze_row = breese_line
                for wbreeze_line in WBreeze_line.values():
                    if wbreeze_line != None:
                        wbreeze_row = wbreeze_line
                for excost_line in Excost_line.values():
                    if excost_line != None:
                        excost_row = excost_line
                for out_line in Out_line.values():
                    if out_line != None:
                        out_row = int(out_line)
                for in_lien in In_lien.values():
                    if in_lien != None:
                        in_row = int(in_lien)
                for m_line in M_line.values():
                    if m_line != None:
                        mass_row = int(m_line)

                print ("reed good")
            print ("step 2", Pulse_line)
            print (sheet_Lines_V)
            workbook.close()
        except:
            print("reader")
###-----------------------------------------------------------------------------------------------
            
            
    def self_data_read(self):

        try:
            read_lines = open("data_mail.txt", "r",encoding = 'utf-8')
            data_mail= read_lines.readlines()
            first_name = self.root.ids.screen3.ids.first_name
            birthday = self.root.ids.screen3.ids.birthday
            sex = self.root.ids.screen3.ids.sex
            growth = self.root.ids.screen3.ids.growth
            number_self = self.root.ids.screen3.ids.number_self
            number_doc = self.root.ids.screen3.ids.number_doc
            number_doc2 = self.root.ids.screen3.ids.number_doc2
            mail_to = self.root.ids.screen3.ids.mail_to

            global data_list
            
            
            

            name = data_mail[1].split("\n")
            if name[0] != '':
                first_name.text=name[0]
                data_list[1] = name[0]

            name = data_mail[2].split("\n")
            if name[0] != '':
                birthday.text=name[0]
                data_list[2] = name[0]

            name = data_mail[3].split("\n")
            if name[0] != '':
                sex.text=name[0]
                data_list[3] = name[0]

            name = data_mail[4].split("\n")
            if name[0] != '':
                growth.text=name[0]
                data_list[4] = name[0]

            name = data_mail[5].split("\n")
            if name[0] != '':
                number_self.text=name[0]
                data_list[5] = name[0]

            name = data_mail[6].split("\n")
            if name[0] != '':
                number_doc.text=name[0]
                data_list[6] = name[0]

            name = data_mail[7].split("\n")
            if name[0] != '':
                number_doc2.text=name[0]
                data_list[7] = name[0]

            name = data_mail[8].split("\n")
            if name[0] != '':
                mail_to.text=name[0]
                data_list[8] = name[0]

        except:
            print ("self_data_read")
        pass

    def self_data_seve(self, line, data):
        try:
            my_file1 = open("data_mail.txt", 'a',encoding = 'utf-8')
            if line == 1:
                if data == "":
                    name = my_file1.write(f'\n')
                else:
                    name = my_file1.write(f'{data}\n')
            if line == 2:
                if data == "":
                    name = my_file1.write(f'\n')
                else:
                    name = my_file1.write(f'{data}\n')
            if line == 3:
                if data == "":
                    name = my_file1.write(f'\n')
                else:
                    name = my_file1.write(f'{data}\n')
            if line == 4:
                if data == "":
                    name = my_file1.write(f'\n')
                else:
                    name = my_file1.write(f'{data}\n')
            if line == 5:
                if data == "":
                    name = my_file1.write(f'\n')
                else:
                    name = my_file1.write(f'{data}\n')
            if line == 6:
                if data == "":
                    name = my_file1.write(f'\n')
                else:
                    name = my_file1.write(f'{data}\n')
            if line == 7:
                if data == "":
                    name = my_file1.write(f'\n')
                else:
                    name = my_file1.write(f'{data}\n')
            if line == 8:
                if data == "":
                    name = my_file1.write(f'\n')
                else:
                    name = my_file1.write(f'{data}\n')
            my_file1.close()
        except: 
            print ("self_data_seve") 
            pass
    def starter(self):
        my_file = open("data_mail.txt", 'w',encoding = 'utf-8')
        my_file.write(f'data_mail\n')
        my_file.close()
#------------------------------------------------------------------------------------------------------

    def data_base(self, line, data):
        global block_list
        block_list[line]= data
    def first_screen(self):
        try: # 1 screen label 1 - 4
            text_label3 = self.root.ids.screen1.ids.text_label3
            text_label3.text = f"{up} / {douwn}"
            text_label3.font_size = '60dp'
            text_label4 = self.root.ids.screen1.ids.text_label4
            text_label4.text = f"пульс {pulse} уд. / мин"
            text_label4.font_size = '20dp'
        except:
            print ("first_screen")

    def graph_pulse(self ):
        print ("graph1")
        graph1= self.root.ids.pulse.ids.bargraph.ids.graph_1
        graph = ObjectProperty()
        xmax_c = 1
        
        counter_x = 0
        plot = MeshLinePlot(color=[1, 0, 0, 1])
        #plot1 =plot.points = [(x, sin(x / 10.)) for x in range(0, 101)]
        for x in Up_line.values():
            if x != None:
                counter_x +=1
                xmax_c +=1
                x = int(x) /200
                plot.points  =plot.points = [(0,0.4),(counter_x,x),]
        #plot.points = [(0,0), (1,1),(2,0),(3,1)]
        # Recherche des widgets
        print(plot.points)
        graph1.xmax = xmax_c
        graph1.add_plot(plot)
        

    def pulse_screen_input(self):

        pulse_up = self.root.ids.pulse.ids.pulse_up
        text_input2 = self.root.ids.pulse.ids.pulse_down
        text_input3 = self.root.ids.pulse.ids.pulse_self
        text_input4 = self.root.ids.pulse.ids.mass
        print (up_row)
        print (douwn_row)
        print (pulse_row)
        print (mass_row)
        pulse_up.text = f"{up_row}"
        text_input2.text  = f"{douwn_row}"
        text_input3.text  = f"{pulse_row}"
        text_input4.text  = f"{mass_row}"       
        
    def pulse_screen(self,up,down,pulse,mass):
        global up_row
        global douwn_row
        global pulse_row
        global mass_row
        global up_row
        global douwn_row
        global pulse_row
        global mass_row
        try:           
            up_row = up
            douwn_row= down
            pulse_row= pulse
            mass_row= mass
        except:
            print ("pulse_screen")
    def graph_breese(self ):
        print ("graph3")
        graph1= self.root.ids.breese.ids.bargraph.ids.graph_1
        graph = ObjectProperty()
        xmax_c = 1
        
        counter_x = 1
        plot = MeshLinePlot(color=[1, 0, 0, 1])
        #plot1 =plot.points = [(x, sin(x / 10.)) for x in range(0, 101)]
        for x in Up_line.values():
            if x != None:
                counter_x +=1
                xmax_c +=1
                x = int(x) /200
                plot.points  =plot.points = [(0,0.4),(counter_x,x),]
        #plot.points = [(0,0), (1,1),(2,0),(3,1)]
        # Recherche des widgets
        print(plot.points)
        graph1.xmax = xmax_c
        graph1.add_plot(plot)
    def breese_screen(self):
        global breeze_row
        global wbreeze_row
        global excost_row
        try:
           breeze_row = block_list[5]
           wbreeze_row =  block_list[6]
           excost_row = block_list[7]
        except:
            print ("breese_screen")
    def result_screen(self, num):
        graph1= self.root.ids.breese.ids.bargraph.ids.graph_1
        graph = ObjectProperty()



        if num == 1:
            xmax_c = 1
            counter_x = 0
            plot = MeshLinePlot(color=[1, 0, 0, 1])
            for x in Pulse_line.values():
                if x != None:
                    counter_x +=1
                    xmax_c +=1
                    x = int(x) /200
                    plot.points  =plot.points = [(0,0.4),(counter_x,x),]
                print(plot.points)
            graph1.xmax = xmax_c
            graph1.add_plot(plot)
            print(1,"puls")
        if num == 2:
            xmax_c = 1
            counter_x = 0
            plot = MeshLinePlot(color=[1, 0, 0, 1])
            for x in Up_line.values():
                if x != None:
                    counter_x +=1
                    xmax_c +=1
                    x = int(x) /200
                    plot.points  =plot.points = [(0,0.4),(counter_x,x),]
                print(plot.points)
            graph1.xmax = xmax_c
            graph1.add_plot(plot)
            print(2,"up")
        if num == 3:
            xmax_c = 1
            counter_x = 0
            plot = MeshLinePlot(color=[1, 0, 0, 1])
            for x in M_line.values():
                if x != None:
                    counter_x +=1
                    xmax_c +=1
                    x = int(x) /200
                    plot.points  =plot.points = [(0,0.4),(counter_x,x),]
            print(plot.points)
            graph1.xmax = xmax_c
            graph1.add_plot(plot)
            print(3, "mass")

    def secondscreen(self, num, values):
        if num == 1:
            print (values)
        if num == 2:
            print (values)

    def stat(self):
        print (block_list)
        print (value_line)
        print (Time_line)
        print (Drags_line)
        print (Doze_line)
        print (Rade_line)
        print (Up_line)
        print (Down_line)
        print (Pulse_line)
        print (Breese_line)
        print (WBreeze_line)
        print (Excost_line)
        print (Out_line)
        print (In_lien)
        print (M_line)
        print (W_line)
        print(v_row)
        print (up_row, douwn_row)















laters = {1:'A',2:'B',3:'C',4:'D',5:'E',6:'F',7:'G',8:'H',9:'I',10:'J',11:'K',12:'L',13:'M',14:"V",15:"W",16:"U"}

value_line= ["1","2"]
Time_line= {}
Drags_line= {}
Doze_line= {}
Rade_line= {}
Up_line= {}
Down_line={}
Pulse_line= {}
Breese_line= {}
WBreeze_line={}
Excost_line= {}
Out_line= {}
In_lien= {}
V_line={1:10}
W_line= {1:5}
U_line= {1:2}
M_line= {}

up_row=0
douwn_row=0
pulse_row=0
mass_row=0
time_row=0
breeze_row=""
wbreeze_row=""
excost_row=""
out_row=0
in_row=0

v_row =1

expenses = {1:"Вреемя",
            2:"Лекарство",
            3:"Дозировка",
            4:"Количество приёма", 
            5:"Верхнее давление", 
            6:"Нижнее давление", 
            7:"Пульс", 
            8:"Одышка", 
            9:"Отеки", 
            10:"Утомляемость",
            11:"Вышло из организма", 
            12:"Выпито", 

            13:"Масса",
            14:v_row,
            15:4,
            16:2}




counter_drags = 0

Text_info = {
    "0":"0",
    1: "Дневник пациента с сердечной недостаточностью(СН) разработан с целью помочь Вам принимать активное участие в лечении вашего заболевания вместе с лечащим врачом. "
    ' Вы сами сможете контролировать ключевые показатели состояния здоровья и в случае необходимости сообщить об изменениях лечащему врачу. '
    ' Кроме того, дневник поиожет Вам разобраться в сущности заболевания и лечебных методах, а также понять важность изменения образа жизни.\n'
    ' Ежедневные записи Ваших измерений будут очень полезны для обсуждения с врачом при последующих посещениях лечебного учереждения.\n '
    ' Вы можете узнать про сердечную недостаточность на официальном образовательном веб-сайте Белорусской ассоциации сердечной недостаточности:'
    'www.heartfailure.by',
    2: "Сердечная недостаточность - это нарушение функции сердца, при котором оно не может перекачивать кровь в достаточном объеме, в органах и тканях."
    ' В результате клетки тела получают недостаточное количество питательных веществ, испытывают кислородное голодание.\n'
    ' СН сопровождается комплексом симптомов: \n'
    ' - одышка;\n'
    ' - усталость;\n'
    ' - отеки(задержка жидкости).\n'
    ' На начальных стадиях симптомы хронической сердечной недостаточности (ХСН) возникают только во время физических нагрузок. '
    ' Со временем эти симптомы усиливаются, начинают беспокоить не только во время физической работы, но и в состоянии покоя. ',
    3: "- Медикаментозное лечение;\n"
    '- Хирургическое (по показаниям);\n'
    '- Электрофизиологические методы терапии (имплантация электрокардиостимуляторов, ресинхронизируещих устройств, дефибриляторов; денервация почечных артерий);\n'
    '- Эндоваскулярная имплантация мезенхимальных стволовых клеток;\n'
    '- Психологическая реабилитация;\n'
    '- Организация школ пациентов с сердечной недостаточностью (формирование мотивации к здоровому образу жизни);\n'
    '- Проведение ежегодных дней знаний о сердечной недостаточности. ',
    4: " Узнать, если у Вас проблемы с весом можно с помощью индекса массы тела (ИМТ), который рассчитывается по формуле:\n"
    '                ИМТ = масса тела (кг)/(рост,м)*(рост,м)\n '
    ' Наличие ожирения или избыточного веса ухудшает прогноз пациента с ХСН, поэтому при ИМТ более 25кг/м.кв. требуется принятие специальных мер и ограничения калорийности питания.\n'
    '                 ИМТ, кг/м.кв.\n' 
    ' Норма 18,5-24,9\n'
    ' Избыточный вес 25,0-29,0\n'
    ' Ожирение >= 30,0\n'
    ' Если у Вас увеличился вес более 2 кг в течении 1-3 дней при соблюдении вашего обычного рациона питания, то необходимо проконсультироваться с врачом, так как это может быть в следствие задержки жидкости в организме! ',
    5: "- Ограничьте употребление поваренной соли: суточное количество соли не должно превышать 3-5 г. "
    'Пищу лучше не подсаливать, достаточно той соли, которая имеется в продуктах.\n'
    '- Исключите газированные напитки.\n'
    '- Избегайте "фаст-фуда".\n'
    '- Ешьте рыбу 1-2 раза в неделю, в том числе рыбу жирных сортов.\n'
    '- Включите в ежедневный рацион по 200 г. фруктов и овощей.\n'
    '- Желательно употребление 30-45 г. в сутки пищевых волокон, особенно из цельнозерновых продуктов. ', 
    6: "Допускается прием до 7 бокалов вина в неделю (1 бокал в сутки).\n"
    ' Большее потребление алкоголя может индуцировать развитие кардиомиопатии токсического генеза.'
    ' В таком случае, прием алкоголя категорически противопоказан.',
    7: "  Курение оказывает неблагоприятное влияние на здоровье и ухудшает прогноз при сердечно-сосудистой патологии.\n"
    ' Курение усиливает эффект таких факторов риска, как возраст, мужской сахарный диабет.\n'
    ' Прекращение курения - одна из самых эффективных мер профилактики сердечно-сосудистых заболеваний.\n'
    ' Избегайте пассивного курения ',
    8: "Малоподвижный образ жизни и низкая физическая активность увеличивает риск раннего развития и прогрессирования ХСН\n"
    '  Не менее 30 минут в день посвящайте упражнениям легкой (пешая прогулка) и умеренной интенсивности (быстрая ходьба). Одним из вариантов физической нагрузки является скандинавская ходьба. '
    ' Выбор режима физической нагрузки зависит от Вашего состояния, оцененного вашим лечащим врачом.',
    9: " Поддерживайте уровень АД меньше 140/90 мм рт.ст.\n"
    ' Как правильно измерить АД в домашних условиях:\n'
    ' - перед измерением необходимо отдохнуть 3-5 минут;\n'
    ' - измеряйте АД в положении сидя, с опорой для спины и рук, поставьте ноги на пол, расположив руку на уровне сердца. Во время измерения не следует разговаривать, в т.ч. и по телефону;\n'
    ' - в идеале нужно производить 2 измерения с интервалом 1-2 мин., 1 раз утром и 1 раз вечером\n'
    ' Как измерить ЧСС в домашних условиях:\n'
    ' Оптимальное ЧСС 55-60 уд/мин.\n'
    ' - посидите спокойно не менее 5 минут;\n'
    ' - пульс удобно прощупывать на запястье руки у основания большого пальца. Удобнее это делать 4 пальцами, при этом 5 палец должен использоваться как опора.\n'
    ' - подсчитайте пульс в течение 60 секунд',    
    10: "  Уровень глюкозы крови у пациентов с ХСН не должен превышать 5 ммоль/л, уровень гликированного гемоглобина 5,6%.",
    11: "  Наличие отеков можно определить простым нажатием пальца в течении нескольких секунд на кожу в области лодыжек.  "
    ' Если нет отека, то после отнятия пальца ткани моментально расправляются. '
    ' Если есть отек, то остается ямка на 1-2 минуты, затем исчезает.',
    12: "Необходимым условием эффективной профилактики СН и ее прогрессирования являются хорошее взаимопонимание между врачом и пациентом.\n"
    ' Эффективность лекарственного средства зависит от его правильного и регулярного применения.'
    ' Не секрет, что пациенты часо не придерживаются предписанной схемы лечения по разным причинам, вследствие чего наступает обострение(декомпенсация) сердечной недостаточности.\n '
    ' Строгое соблюдение всех рекомендацией врача по проводимой терапии и принципов здорового образа жизни - залог уменьшения симптомов сердечной недостаточности, улучшения качества жизни пациента и предотвращения прогрессирования заболеваний!',
    13: "Незамедлительно проинформируйте своего лечащего врача, если Вы испытываете:\n "
    ' - усиление одышки;\n'
    ' - уменьшение переносимости физических нагрузок;\n '
    ' - частые пробуждения ночью из-за одышки и потребности в большем количестве подушек для обеспечения комфорта;\n'
    ' - стойкое учащенное сердцебиение;\n'
    ' - нарушение ритма сердечных сокращений;\n '
    ' - быстрое увеличение массы тела;\n'
    ' Следует вызвать скорую помощь в случае:\n'
    ' - постоянной боли в груди, от которой не избавляет прием нитроглицерина;\n'
    ' - тяжелой и стойкой одышки;\n'
    ' - слабости или обморока;',
}

Text_info1 = {
    "0":"0",    
    1:"АСК",
    2:'Аккузид',
    3:'Аккупро',
    4:'Амиодарон',
    5:'Амиокордин',
    6:'Амлесса',
    7:'Амлодин',
    8:'Амлодипин',
    9:'Амлотензин',
    10:'Амприлан',
    11:'Арифон ретард',
    12:'Асомекс',
    13:'Аспикард',
    14:'Аторвастатин',
    15:'Аторис',
    16:'Аудитор',
    17:'Ацетилсалициловая кислота',
    18:'Берлиприл',
    19:'Берлиприл Плюс',
    20:'Беталок',
    21:"Бикард",
    22:'Бисопролол',
    23:'Бравадин',
    24:'Валз',
    25:'Валз Н',
    26:'Валзан',
    27:'Валзан Н',
    28:'Валсартан',
    29:'Вальсакор',
    30:'Вальсакор Н',
    31:'Вальсакор Нд',
    32:'Варфарекс',
    33:'Варфарин',
    34:'Васкопин',
    35:'Верапамил',
    36:'Верошпирон',
    37:'Гидрохлортиазид',
    38:'Гипотиазид',
    39:'Даприл',
    40:'Дигоксин',
    41:"Диласидом",
    42:'Дилатренд',
    43:'Дилтиазем',
    44:'Дилтиазем ретард',
    45:'Диован',
    46:'Диротон',
    47:'Диувер',
    48:'Ивабрадин',
    49:'Изо-мик лонг',
    50:'Индалонг',
    51:'Индап',
    52:'Индапамид',
    53:'Индапафон',
    54:'Индапен',
    55:'Индапен ретард',
    56:'Индопрес',
    57:'Инспра',
    58:'Калчек',
    59:'Кандесартан',
    60:'Карведил',
    61:"Карведилол",
    62:'Карвелэнд',
    63:'Кардивас',
    64:'Кардикет',
    65:'Кардилопин',
    66:'Кардиомагнум',
    67:'Квинаприл',
    68:'Ко-Диован',
    69:'Ко-амлесса',
    70:'Ко-валсартан',
    71:'Ко-ренитек',
    72:'Конкор',
    73:'Конкор кор',
    74:'Кораксан',
    75:'Кордарон',
    76:'Кордафлекс',
    77:'Кордипин XL',
    78:'Кордипин ретард',
    79:'Коронал',
    80:'Крестор',
    81:"Ксарелто",
    82:'Леркамен',
    83:'Леркадипин',
    84:'Лизиноприл',
    85:'Лизиноприл плюс',
    86:'Лизитар',
    87:'Лизоретик',
    88:'Лизорил',
    89:'Липразид',
    90:'Липримар',
    91:'Ловастатин',
    92:'Лозартан',
    93:'Лориста',
    94:'Магнекард',
    95:'Мертенил',
    96:'Метокард',
    97:'Метопролол',
    98:'Молсидомин',
    99:'Небиволол',
    100:'Небивомед',
    101:"Небилет",
    102:'Небилет плюс',
    103:'Нитроглицерин',
    104:'Нитрокор',
    105:'Нитроминт',
    106:'Нитросорбид',
    107:'Нитроспрей',
    108:'Нифедипин',
    109:'Нифекард XL',
    110:'Норваск',
    111:'Нормодипин',
    112:'Омега 3',
    113:'Памид',
    114:'Паралель',
    115:'Периндоприл',
    116:'Периндоприл плюс',
    117:'Периндоприл+амлодипин',
    118:'Периндоприл+амлодипин+индапамид',
    119:'Полприл',
    120:'Парадакса',
    121:"Предуктал",
    122:'Принесса',
    123:'Престариум',
    124:'Рамилонг',
    125:'Рамиприл',
    126:'Рамприл',
    127:'Равел СР',
    128:'Ревелол XL',
    129:'Ренитек',
    130:'Розувастатин',
    131:'Розукард',
    132:'Розулип',
    133:'Розутатин',
    134:'Ромазик',
    135:'Сентор',
    136:'Сиднофарм',
    137:'Синатор',
    138:'Спироналктон',
    139:'Стамло',
    140:'Сустонит',
    141:"Тиотриазолин",
    142:'Торасемид',
    143:'Три-зидин М',
    144:'Трикард',
    145:'Триметазидин',
    146:'Трипликсам',
    147:'Тритаце',
    148:'Трован',
    149:'Тромбо асс',
    150:'Тулип',
    151:'Фозикард',
    152:'Фозикард Н',
    153:'Фозиноприл',
    154:'Фуросемид',
    155:'Хартил',
    156:'Хинаприл',
    157:'Эгилок',
    158:'Эгилок СР',
    159:'Экваприл',
    160:'Экватор',
    161:"Эналаприл",
    162:'Эналаприл HL',
    163:'Эналаприл Н',
    164:'Энам',
    165:'Энамед Н',
    166:'Энап',
    167:'Энап HL',
    168:'Энап Н',
    169:'Энаприл',
    170:'Энаприл-НТ',
    171:'Энаренал',
    172:'Энтресто (сакубитрил/валсартан)',
    173:'Эплеренон',
    174:'липромак',
    175:'нитрогранулонг',

}


block_list= {
    "0":"0",
    1:"",
    2:"",
    3:"",
    4:"",
    5:"",
    6:"",
    7:"",
    8:"",
    9:"9",
    "10":"10",
    "11":"11",

    }
data_list= {
    "0":"0",
    1:"",
    2:"",
    3:"",
    4:"",
    5:"",
    6:"",
    7:"",
    8:"",
    9:"9",
    "10":"10",
    "11":"11",

    }
drags_list= {}
drags_doz= {}
drags_rate= {}


lines_reader= 0
lines_counter= 0
lines_last= 0



Dizar= ["Другое", "0,2 мг", '0,25мг','0,4мг','1,25мг','1,5мг','10мг','100мг','10мг/2,5/10мг','10мг/2,5мг/5мг','150мг','15мг','160мг','2мг','2,5мг','20мг','200мг','30мг','35мг','4мг','4/10/1,25мг','4/10мг','4/5мг','4/5/1,25мг','40мг','5','50','5мг/1,25мг/10мг','50мг','5мг/1,25мг/10мг','5мг/1,25мг/5мг','6мг','6,25мг','60мг','7,5мг','8мг','8/10/2,5мг','8/2,5мг','8/10мг','8/5/2,5мг','8/5мг','90мг','97/103мг',       


        ]
Days= ["Другое", "1 раз в день",'2 раза в день','3 раза в день','4 раза в день','5 раз в день',
        


        ]







        
MainApp().run()